#!/usr/bin/env python3
"""
Build two filled ingestion workbooks for an *insurance reserving* SAS -> Snowflake
migration (an AS-IS migration: same tables/columns in both systems).

  - Reserving_SAS_Legacy_Filled.xlsx       (import as the LEGACY system)
  - Reserving_Snowflake_Target_Filled.xlsx (import as the TARGET system)

Both load the committed template (preserving its dropdowns/structure) and fill
TABLE_1..TABLE_8 plus the MASTER registry + table connections, exactly like
scripts/build_sample.py. Import each file separately into the SAME canvas, choosing
Legacy for the SAS file and Target for the Snowflake file.

Design notes
------------
- 8 tables, ~12 columns each, mostly numerical with a few identifiers, categoricals
  and dates — a realistic actuarial loss-reserving model.
- Proper per-system data types:
    SAS      -> NUM (numbers & SAS dates) and CHAR (text)
    Snowflake-> NUMBER(p,s)/FLOAT for numbers, VARCHAR(n) for text, DATE for dates
  (So NUM<->NUMBER and SAS-date-NUM<->DATE legitimately read as type changes, while
   text and true ratios stay equivalent — useful signal for the status dashboard.)
- Table names and column names are slightly different across the two systems.
- Stats carry slight reconciliation drift between systems (row counts, sums, means,
  null/unique counts), as asked.

Run:
    /tmp/tplvenv/bin/python scripts/build_reserving_samples.py
"""

import hashlib
import os
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
TPL_DIR = os.path.abspath(os.path.join(HERE, "..", "public", "templates"))
TEMPLATE = os.path.join(TPL_DIR, "Lineage_Canvas_Template.xlsx")
OUT_SAS = os.path.join(TPL_DIR, "Reserving_SAS_Legacy_Filled.xlsx")
OUT_SF = os.path.join(TPL_DIR, "Reserving_Snowflake_Target_Filled.xlsx")

# ---------------------------------------------------------------------------
# Template-writing helpers (copied from scripts/build_sample.py)
# ---------------------------------------------------------------------------

def set_kv(ws, key, value):
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == key:
            ws.cell(row=row[0].row, column=2, value=value)
            return
    raise KeyError(f"{key} not found on {ws.title}")


def header_row(ws, first_cell):
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == first_cell:
            return row[0].row
    raise KeyError(f"header {first_cell} not found on {ws.title}")


def headers_at(ws, hrow):
    out, c = {}, 1
    while True:
        v = ws.cell(row=hrow, column=c).value
        if v is None or str(v).strip() == "":
            break
        out[str(v).strip()] = c
        c += 1
    return out


def fill_grid(ws, first_cell, rows):
    hrow = header_row(ws, first_cell)
    cols = headers_at(ws, hrow)
    for i, record in enumerate(rows):
        r = hrow + 1 + i
        for key, val in record.items():
            if key in cols:
                ws.cell(row=r, column=cols[key], value=val)


def fill_table(ws, meta, columns):
    for k, v in meta.items():
        set_kv(ws, k, v)
    fill_grid(ws, "column_name", columns)


def fill_registry(ws, mapping):
    hrow = header_row(ws, "sheet_name")
    r = hrow + 1
    while True:
        sheet = ws.cell(row=r, column=1).value
        if sheet is None or str(sheet).strip() == "":
            break
        name = mapping.get(str(sheet).strip())
        if name:
            ws.cell(row=r, column=2, value=name)
        r += 1


# ---------------------------------------------------------------------------
# Deterministic pseudo-stats (stable across runs; no RNG state)
# ---------------------------------------------------------------------------

def seed_of(*parts):
    h = hashlib.md5("|".join(str(p) for p in parts).encode()).hexdigest()
    return int(h[:8], 16)


def jitter(seed, lo, hi):
    return lo + (seed % 100003) / 100003 * (hi - lo)


def base_mean(name):
    n = name.upper()
    table = [
        ("LIMIT", 285000.0), ("DEDUCT", 1450.0), ("RETENTION", 25000.0),
        ("PREMIUM", 1840.0), ("PREM", 1840.0), ("EXPOSURE", 12.5),
        ("IBNR", 9200.0), ("ULTIMATE", 38500.0), ("ULT", 38500.0),
        ("INCURRED", 27600.0), ("INCD", 27600.0), ("CUMULATIVE", 220000.0),
        ("CUM", 220000.0), ("PAID", 18900.0), ("PD_", 18900.0),
        ("CASE", 11200.0), ("MARGIN", 5300.0), ("RESERVE", 14300.0),
        ("RSV", 14300.0), ("NET", 16500.0), ("ALAE", 2200.0),
        ("RECOV", 1300.0),
    ]
    for kw, v in table:
        if kw in n:
            return v
    return 9500.0


def base_ratio(name):
    n = name.upper()
    table = [
        ("LOSS_RATIO", 0.642), ("ELR", 0.642), ("EXPECTED_LOSS", 0.642),
        ("LDF", 1.185), ("DEVELOPMENT_FACTOR", 1.185),
        ("ILF", 1.32), ("INCREASED_LIMIT", 1.32),
        ("ADEQ", 1.04), ("ON_LEVEL", 1.07), ("ONLVL", 1.07),
        ("EARN", 0.58),
    ]
    for kw, v in table:
        if kw in n:
            return v
    return 1.0


def cat_values(name):
    n = name.upper()
    if "LINE_OF_BUSINESS" in n or n == "LOB_CD":
        return (["AUTO", "HOME", "GL", "WC", "CPP", "BOP"], 6)
    if "CLAIM_STATUS" in n or "CLM_STAT" in n:
        return (["OPEN", "CLOSED", "REOPENED", "DENIED"], 4)
    if "POLICY_STATUS" in n or "POL_STAT" in n:
        return (["INFORCE", "CANCELLED", "EXPIRED", "NONRENEWED"], 4)
    if "RESERVE_STATUS" in n or "RSV_STAT" in n:
        return (["ACTIVE", "CLOSED", "REVIEW"], 3)
    if "CAUSE" in n:
        return (["FIRE", "WATER", "THEFT", "COLLISION", "WIND", "HAIL", "LIABILITY"], 7)
    if "STATE" in n or n == "ST_CD":
        return (["CA", "TX", "NY", "FL", "IL", "PA", "OH", "GA", "NC", "NJ"], 50)
    if "TRANSACTION_TYPE" in n or "TXN_TYP" in n:
        return (["PAYMENT", "RESERVE", "RECOVERY", "EXPENSE"], 4)
    if "PERIL" in n:
        return (["FIRE", "WIND", "HAIL", "WATER", "THEFT", "FLOOD"], 6)
    if "COVERAGE_CODE" in n or "COV_CD" in n:
        return (["DWELL", "LIAB", "COLL", "COMP", "MEDPAY", "PD"], 6)
    if "FINANCIAL_PERIOD" in n or "FIN_PERIOD" in n:
        return (["2023-01", "2023-06", "2023-12", "2024-06", "2024-12"], 24)
    return (["A", "B", "C", "D"], 4)


def stats_for(role, name, rows, sysseed):
    """Return a dict of the stat cell values for one column on one system."""
    s = seed_of(name, sysseed)
    out = {}
    if role in ("id", "catid"):
        out["null_count"] = 0
        out["unique_count"] = max(0, rows - (s % 3))   # near-unique
        return out
    if role == "cat":
        vals, uniq = cat_values(name)
        out["null_count"] = s % 80
        out["unique_count"] = uniq + (s % 3 - 1)
        out["uniques"] = ",".join(vals[:6])
        return out
    if role == "date":
        out["null_count"] = (s % 60)
        out["min_value"] = "2015-01-05"
        out["max_value"] = "2024-12-28"
        return out
    if role == "int":
        n = name.upper()
        if "YEAR" in n or n.endswith("_YR"):
            out.update(min_value="2015", max_value="2024", unique_count=10, null_count=0)
        elif "MONTH" in n or "DEV_MO" in n:
            out.update(min_value="0", max_value="120", mean_value=round(jitter(s, 48, 60), 1),
                       unique_count=11, null_count=0)
        elif "TERM" in n:
            out.update(min_value="6", max_value="12", mean_value=round(jitter(s, 11.0, 11.8), 2),
                       unique_count=3, null_count=0)
        else:  # a count
            mean = round(jitter(s, 3, 42), 2)
            nn = rows - (s % 50)
            out.update(min_value="0", max_value=str(int(mean * 9 + 30)),
                       mean_value=mean, sum_value=round(mean * nn, 2),
                       null_count=s % 50)
        return out
    if role == "float":
        base = base_ratio(name)
        out.update(
            min_value=round(base * 0.45, 4), max_value=round(base * 1.8, 4),
            mean_value=round(base * (1 + (s % 200 - 100) / 100000), 4),
            stddev_value=round(base * 0.22, 4), null_count=s % 120,
        )
        return out
    # role == "num" (monetary amount)
    base = base_mean(name)
    nulls = s % 600
    nn = max(1, rows - nulls)
    mean = round(base * (1 + (s % 300 - 150) / 100000), 2)
    out.update(
        null_count=nulls,
        min_value=0,
        max_value=round(mean * jitter(s + 7, 9, 24), 2),
        mean_value=mean,
        stddev_value=round(mean * 0.7, 2),
        sum_value=round(mean * nn, 2),
    )
    return out


# ---------------------------------------------------------------------------
# Per-system data-type mapping
# ---------------------------------------------------------------------------

def sas_type(role):
    return "CHAR" if role in ("cat", "catid") else "NUM"


def sf_type(role):
    return {
        "id": "NUMBER(38,0)", "int": "NUMBER(10,0)", "num": "NUMBER(18,2)",
        "float": "FLOAT", "date": "DATE", "cat": "VARCHAR", "catid": "VARCHAR",
    }[role]


def sf_precision(role, name):
    n = name.upper()
    if role == "id":
        return 38
    if role == "num":
        return 18
    if role == "int":
        if "YEAR" in n or n.endswith("_YR"):
            return 4
        if "TERM" in n:
            return 2
        return 10
    return None


def text_len(name):
    n = name.upper()
    if "NUMBER" in n or "NBR" in n:
        return 30
    if "STATE" in n or n == "ST_CD":
        return 2
    if "PERIOD" in n:
        return 7
    if "CODE" in n or n.endswith("_CD"):
        return 20
    if "STATUS" in n or "STAT" in n or "TYPE" in n or "TYP" in n or "CAUSE" in n:
        return 30
    return 40


# ---------------------------------------------------------------------------
# The reserving model: 8 tables. Each column = (sas_name, sf_name, role).
# roles: id (numeric id), catid (text id), int, num (amount), float (ratio), cat, date
# ---------------------------------------------------------------------------

T = [
    dict(
        sheet="TABLE_1", key="t1", sas="CLM_HDR", sf="CLAIM_HEADER",
        sas_ns="RESVCLM", sf_ns="ACTUARIAL.CLAIMS", domain="Claims",
        desc="Claim header — one row per claim (reserving view).",
        grain="one row per claim", refresh="DAILY",
        sas_rows=412_540, sf_rows=412_488, key_sas="CLM_ID", key_sf="CLAIM_ID",
        cols=[
            ("CLM_ID", "CLAIM_ID", "id"),
            ("CLM_NBR", "CLAIM_NUMBER", "catid"),
            ("POL_ID", "POLICY_ID", "id"),
            ("LOB_CD", "LINE_OF_BUSINESS_CODE", "cat"),
            ("CLM_STAT", "CLAIM_STATUS", "cat"),
            ("CAUSE_CD", "CAUSE_OF_LOSS_CODE", "cat"),
            ("ACDT_DT", "ACCIDENT_DATE", "date"),
            ("RPT_DT", "REPORT_DATE", "date"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("INIT_RSV", "INITIAL_RESERVE_AMOUNT", "num"),
            ("PD_TO_DT", "PAID_TO_DATE_AMOUNT", "num"),
            ("TOT_INCD", "TOTAL_INCURRED_AMOUNT", "num"),
        ],
    ),
    dict(
        sheet="TABLE_2", key="t2", sas="CLM_TXN", sf="CLAIM_TRANSACTION",
        sas_ns="RESVCLM", sf_ns="ACTUARIAL.CLAIMS", domain="Claims",
        desc="Claim financial transactions — payments and reserve movements.",
        grain="one row per claim financial transaction", refresh="DAILY",
        sas_rows=2_948_110, sf_rows=2_947_690, key_sas="TXN_ID", key_sf="TRANSACTION_ID",
        cols=[
            ("TXN_ID", "TRANSACTION_ID", "id"),
            ("CLM_ID", "CLAIM_ID", "id"),
            ("TXN_TYP", "TRANSACTION_TYPE", "cat"),
            ("FIN_PERIOD", "FINANCIAL_PERIOD", "cat"),
            ("TXN_DT", "TRANSACTION_DATE", "date"),
            ("DEV_MO", "DEVELOPMENT_MONTH", "int"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("PD_LOSS", "PAID_LOSS_AMOUNT", "num"),
            ("PD_ALAE", "PAID_ALAE_AMOUNT", "num"),
            ("RSV_CHG", "RESERVE_CHANGE_AMOUNT", "num"),
            ("RECOV_AMT", "RECOVERY_AMOUNT", "num"),
            ("NET_AMT", "NET_AMOUNT", "num"),
        ],
    ),
    dict(
        sheet="TABLE_3", key="t3", sas="POL_MAST", sf="POLICY_MASTER",
        sas_ns="RESVPOL", sf_ns="ACTUARIAL.POLICY", domain="Policy",
        desc="Policy master — one row per policy term (exposure source).",
        grain="one row per policy term", refresh="DAILY",
        sas_rows=1_503_220, sf_rows=1_503_220, key_sas="POL_ID", key_sf="POLICY_ID",
        cols=[
            ("POL_ID", "POLICY_ID", "id"),
            ("POL_NBR", "POLICY_NUMBER", "catid"),
            ("LOB_CD", "LINE_OF_BUSINESS_CODE", "cat"),
            ("ST_CD", "STATE_CODE", "cat"),
            ("POL_STAT", "POLICY_STATUS", "cat"),
            ("EFF_DT", "EFFECTIVE_DATE", "date"),
            ("EXP_DT", "EXPIRATION_DATE", "date"),
            ("UW_YR", "UNDERWRITING_YEAR", "int"),
            ("TERM_MO", "TERM_MONTHS", "int"),
            ("WRTN_PREM", "WRITTEN_PREMIUM_AMOUNT", "num"),
            ("ERND_PREM", "EARNED_PREMIUM_AMOUNT", "num"),
            ("EXPOSURE", "EXPOSURE_UNITS", "num"),
        ],
    ),
    dict(
        sheet="TABLE_4", key="t4", sas="COV_DTL", sf="COVERAGE_DETAIL",
        sas_ns="RESVPOL", sf_ns="ACTUARIAL.POLICY", domain="Policy",
        desc="Coverage detail — limits, deductibles and coverage premium.",
        grain="one row per coverage per policy", refresh="DAILY",
        sas_rows=5_311_400, sf_rows=5_310_120, key_sas="COV_ID", key_sf="COVERAGE_ID",
        cols=[
            ("COV_ID", "COVERAGE_ID", "id"),
            ("POL_ID", "POLICY_ID", "id"),
            ("COV_CD", "COVERAGE_CODE", "cat"),
            ("PERIL_CD", "PERIL_CODE", "cat"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("LIM_AMT", "LIMIT_AMOUNT", "num"),
            ("DED_AMT", "DEDUCTIBLE_AMOUNT", "num"),
            ("RETN_AMT", "RETENTION_AMOUNT", "num"),
            ("COV_PREM", "COVERAGE_PREMIUM_AMOUNT", "num"),
            ("EXPOSURE", "EXPOSURE_UNITS", "num"),
            ("ILF", "INCREASED_LIMIT_FACTOR", "float"),
            ("EARN_FCT", "EARNING_FACTOR", "float"),
        ],
    ),
    dict(
        sheet="TABLE_5", key="t5", sas="CASE_RSV", sf="CASE_RESERVE",
        sas_ns="RESVACT", sf_ns="ACTUARIAL.RESERVING", domain="Reserving",
        desc="Case reserve valuations over time (per claim per valuation date).",
        grain="one row per claim per valuation date", refresh="MONTHLY",
        sas_rows=3_902_775, sf_rows=3_902_410, key_sas="RSV_ID", key_sf="RESERVE_ID",
        cols=[
            ("RSV_ID", "RESERVE_ID", "id"),
            ("CLM_ID", "CLAIM_ID", "id"),
            ("VAL_DT", "VALUATION_DATE", "date"),
            ("RSV_STAT", "RESERVE_STATUS", "cat"),
            ("DEV_MO", "DEVELOPMENT_MONTH", "int"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("CASE_RSV_AMT", "CASE_RESERVE_AMOUNT", "num"),
            ("IBNR_AMT", "IBNR_AMOUNT", "num"),
            ("PD_TO_DT", "PAID_TO_DATE_AMOUNT", "num"),
            ("INCD_LOSS", "INCURRED_LOSS_AMOUNT", "num"),
            ("ULT_LOSS", "ULTIMATE_LOSS_AMOUNT", "num"),
            ("ADEQ_RATIO", "ADEQUACY_RATIO", "float"),
        ],
    ),
    dict(
        sheet="TABLE_6", key="t6", sas="LOSS_TRI", sf="LOSS_DEVELOPMENT_TRIANGLE",
        sas_ns="RESVACT", sf_ns="ACTUARIAL.RESERVING", domain="Reserving",
        desc="Loss development triangle by accident year and development month.",
        grain="one row per accident year per development month per LOB", refresh="MONTHLY",
        sas_rows=8_640, sf_rows=8_640, key_sas="TRI_ID", key_sf="TRIANGLE_ID",
        cols=[
            ("TRI_ID", "TRIANGLE_ID", "id"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("DEV_MO", "DEVELOPMENT_MONTH", "int"),
            ("LOB_CD", "LINE_OF_BUSINESS_CODE", "cat"),
            ("RPTD_CNT", "REPORTED_CLAIM_COUNT", "int"),
            ("CLSD_CNT", "CLOSED_CLAIM_COUNT", "int"),
            ("CUM_PD_LOSS", "CUMULATIVE_PAID_LOSS_AMOUNT", "num"),
            ("CUM_INCD_LOSS", "CUMULATIVE_INCURRED_LOSS_AMOUNT", "num"),
            ("EARN_PREM", "EARNED_PREMIUM_AMOUNT", "num"),
            ("ULT_LOSS", "SELECTED_ULTIMATE_LOSS_AMOUNT", "num"),
            ("LDF", "LOSS_DEVELOPMENT_FACTOR", "float"),
            ("ELR", "EXPECTED_LOSS_RATIO", "float"),
        ],
    ),
    dict(
        sheet="TABLE_7", key="t7", sas="ERN_PREM", sf="EARNED_PREMIUM",
        sas_ns="RESVPOL", sf_ns="ACTUARIAL.POLICY", domain="Finance",
        desc="Earned premium and exposure by policy and financial period.",
        grain="one row per policy per financial period", refresh="MONTHLY",
        sas_rows=4_120_900, sf_rows=4_120_510, key_sas="EP_ID", key_sf="EARNED_PREMIUM_ID",
        cols=[
            ("EP_ID", "EARNED_PREMIUM_ID", "id"),
            ("POL_ID", "POLICY_ID", "id"),
            ("LOB_CD", "LINE_OF_BUSINESS_CODE", "cat"),
            ("ST_CD", "STATE_CODE", "cat"),
            ("FIN_PERIOD", "FINANCIAL_PERIOD", "cat"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("WRTN_PREM", "WRITTEN_PREMIUM_AMOUNT", "num"),
            ("ERND_PREM", "EARNED_PREMIUM_AMOUNT", "num"),
            ("UNERN_PREM", "UNEARNED_PREMIUM_AMOUNT", "num"),
            ("EXPOSURE", "EARNED_EXPOSURE_UNITS", "num"),
            ("ONLVL_PREM", "ON_LEVEL_PREMIUM_AMOUNT", "num"),
            ("ONLVL_FCT", "ON_LEVEL_FACTOR", "float"),
        ],
    ),
    dict(
        sheet="TABLE_8", key="t8", sas="RSV_SMRY", sf="RESERVE_SUMMARY",
        sas_ns="RESVACT", sf_ns="ACTUARIAL.RESERVING", domain="Reserving",
        desc="Reserve summary / IBNR rollup by accident year and LOB.",
        grain="one row per accident year per LOB per valuation", refresh="MONTHLY",
        sas_rows=10_080, sf_rows=10_080, key_sas="SMRY_ID", key_sf="SUMMARY_ID",
        cols=[
            ("SMRY_ID", "SUMMARY_ID", "id"),
            ("ACDT_YR", "ACCIDENT_YEAR", "int"),
            ("LOB_CD", "LINE_OF_BUSINESS_CODE", "cat"),
            ("VAL_DT", "VALUATION_DATE", "date"),
            ("CLM_CNT", "CLAIM_COUNT", "int"),
            ("EARN_PREM", "EARNED_PREMIUM_AMOUNT", "num"),
            ("PD_LOSS", "PAID_LOSS_AMOUNT", "num"),
            ("CASE_RSV", "CASE_RESERVE_AMOUNT", "num"),
            ("IBNR_AMT", "IBNR_AMOUNT", "num"),
            ("ULT_LOSS", "ULTIMATE_LOSS_AMOUNT", "num"),
            ("RSV_MARGIN", "RESERVE_MARGIN_AMOUNT", "num"),
            ("LOSS_RATIO", "LOSS_RATIO", "float"),
        ],
    ),
]

# Table-level lineage (by table key); names resolved per system.
CONN = [
    ("t3", "t4"), ("t3", "t1"), ("t1", "t2"), ("t1", "t5"), ("t3", "t7"),
    ("t2", "t6"), ("t7", "t6"), ("t6", "t8"), ("t5", "t8"), ("t1", "t8"),
]

NULLABLE_FALSE_ROLES = {"id", "catid"}


def nullable_of(role, name):
    if role in NULLABLE_FALSE_ROLES:
        return "FALSE"
    n = name.upper()
    if role == "date":
        return "FALSE" if ("ACCIDENT" in n or "EFFECTIVE" in n or "ACDT" in n or "EFF" in n) else "TRUE"
    if role == "int":
        return "FALSE" if ("YEAR" in n or n.endswith("_YR") or "MONTH" in n or "DEV_MO" in n) else "TRUE"
    if role == "cat":
        return "FALSE"
    return "TRUE"  # num / float


def build_columns(spec, system):
    """system: 'SAS' or 'SF'. Returns the list of column-row dicts."""
    sysseed = 11 if system == "SAS" else 29
    rows = spec["sas_rows"] if system == "SAS" else spec["sf_rows"]
    out = []
    for sas_name, sf_name, role in spec["cols"]:
        name = sas_name if system == "SAS" else sf_name
        col = {"column_name": name, "nullable": nullable_of(role, name)}
        if system == "SAS":
            col["data_type"] = sas_type(role)
            if role in ("cat", "catid"):
                col["max_length"] = text_len(name)
            if role == "date":
                col["column_definition"] = "SAS date value (DATE9. format)"
        else:
            col["data_type"] = sf_type(role)
            if role in ("cat", "catid"):
                col["max_length"] = text_len(name)
            prec = sf_precision(role, name)
            if prec is not None:
                col["precision"] = prec
        col.update(stats_for(role, name, rows, sysseed))
        out.append(col)
    return out


def build_meta(spec, system):
    name_key = "sas" if system == "SAS" else "sf"
    ns = spec["sas_ns"] if system == "SAS" else spec["sf_ns"]
    rows = spec["sas_rows"] if system == "SAS" else spec["sf_rows"]
    key = spec["key_sas"] if system == "SAS" else spec["key_sf"]
    return {
        "namespace": ns,
        "description": spec["desc"],
        "environment": "PROD",
        "business_domain": spec["domain"],
        "row_count": rows,
        "column_count": len(spec["cols"]),
        "has_primary_key": "TRUE",
        "unique_key_columns": key,
        "grain_description": spec["grain"],
        "refresh_frequency": spec["refresh"],
    }, spec[name_key]


def build_workbook(system, out_path):
    wb = load_workbook(TEMPLATE)
    master = wb["MASTER"]

    name_of = {t["key"]: (t["sas"] if system == "SAS" else t["sf"]) for t in T}
    registry = {t["sheet"]: name_of[t["key"]] for t in T}
    fill_registry(master, registry)

    table_connections = [
        {"from_table": name_of[a], "to_table": name_of[b],
         "description": f"{name_of[a]} feeds {name_of[b]}"}
        for a, b in CONN
    ]
    fill_grid(master, "from_table", table_connections)

    for spec in T:
        meta, _ = build_meta(spec, system)
        fill_table(wb[spec["sheet"]], meta, build_columns(spec, system))

    wb.save(out_path)
    print(f"Wrote {out_path}  ({len(T)} tables)")


def main():
    build_workbook("SAS", OUT_SAS)
    build_workbook("SF", OUT_SF)


if __name__ == "__main__":
    main()
