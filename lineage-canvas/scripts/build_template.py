#!/usr/bin/env python3
"""
Build the Lineage Canvas ingestion template (.xlsx).

The output is a *binary* workbook with native Excel dropdowns (data
validations), which the SheetJS/`xlsx` library used in the app cannot emit.
We therefore author it here with openpyxl and commit the result so the app can
serve it as a static download. The matching importer lives in
src/lib/excelService.ts and parses exactly the structure produced below.

Structure
---------
- INSTRUCTIONS / MASTER are reserved (never imported as tables).
- MASTER has three sections: TABLE REGISTRY, TABLE CONNECTIONS, COLUMN CONNECTIONS.
- The TABLE REGISTRY maps each fixed sheet (TABLE_1..TABLE_15) to a table_name.
  ONLY registry rows that have a table_name are imported (max 15 tables).
- Connection sections reference tables via dropdowns sourced from the registry's
  table_name column, so they can only point at registered tables.
- Each TABLE_n sheet holds table-level metadata (namespace, ...) and a column grid.
  The table's NAME comes from the registry, not the sheet. The project, canvas and
  system (Legacy/Target) are chosen in the app's import screen — one system per import.

Regenerate after changing the structure:

    python3 -m venv /tmp/tplvenv
    /tmp/tplvenv/bin/pip install openpyxl
    /tmp/tplvenv/bin/python scripts/build_template.py

Output: public/templates/Lineage_Canvas_Template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

MAX_TABLES = 15

# ---------------------------------------------------------------------------
# Controlled vocabularies — keep in sync with src/lib/excelService.ts.
# "UNASSIGNED" is the app's sentinel for "no value" (see DetailsPanel.tsx); the
# importer maps it back to undefined.
# ---------------------------------------------------------------------------
ENVIRONMENTS = ["UNASSIGNED", "DEV", "TEST", "UAT", "PROD"]
BOOLEANS = ["TRUE", "FALSE"]
REFRESH = ["UNASSIGNED", "DAILY", "WEEKLY", "MONTHLY", "AD_HOC"]

# Ordered table-level metadata fields: (key, note). table_name is declared in the
# MASTER registry; system (Legacy/Target) is chosen in the app at import time.
TABLE_META_FIELDS = [
    ("namespace", "SAS library or DATABASE.SCHEMA"),
    ("description", "What the table holds"),
    ("environment", "UNASSIGNED / DEV / TEST / UAT / PROD (dropdown)"),
    ("business_domain", "Claims, Policy, Billing, Finance, ..."),
    ("row_count", "Number of rows (number)"),
    ("column_count", "Number of columns (number; left blank = counted automatically)"),
    ("has_primary_key", "TRUE or FALSE (dropdown)"),
    ("unique_key_columns", "Comma-separated column names"),
    ("grain_description", "e.g. one row per policy per term"),
    ("refresh_frequency", "UNASSIGNED / DAILY / WEEKLY / MONTHLY / AD_HOC (dropdown)"),
]

# Ordered column-level fields (header row of the COLUMN METADATA grid).
COLUMN_HEADERS = [
    "column_name", "data_type", "nullable", "max_length", "precision",
    "default_value", "column_definition", "column_computation_formula",
    "null_count", "min_value", "max_value", "unique_count", "uniques",
    "mean_value", "stddev_value", "sum_value",
]

COLUMN_DATA_ROWS = 100          # blank column rows per table sheet
TABLE_TITLE_ROW = 1
META_HEADER_ROW = 2             # "Field | Value | Notes"
META_FIRST_ROW = 3              # first metadata key row
COLUMNS_TITLE_ROW = META_FIRST_ROW + len(TABLE_META_FIELDS) + 1
COLUMNS_HEADER_ROW = COLUMNS_TITLE_ROW + 1
COLUMNS_FIRST_DATA_ROW = COLUMNS_HEADER_ROW + 1

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
TITLE_FONT = Font(bold=True, size=13, color="1F2937")
SECTION_FONT = Font(bold=True, size=11, color="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
KEY_FONT = Font(bold=True, color="111827")
KEY_FILL = PatternFill("solid", fgColor="EFF6FF")
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
INPUT_FILL = PatternFill("solid", fgColor="FFFDF5")
LOCKED_FILL = PatternFill("solid", fgColor="F3F4F6")
EXAMPLE_FONT = Font(italic=True, color="047857")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def list_dv(values, allow_other=False):
    """Inline-list data validation (all our lists are well under the 255-char limit)."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
        showDropDown=False,   # False => arrow IS shown (Excel's inverted flag)
    )
    _set_strictness(dv, allow_other)
    return dv


def range_dv(formula, allow_other=False):
    """List validation whose source is a cell range (e.g. the registry column)."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    _set_strictness(dv, allow_other)
    return dv


def _set_strictness(dv, allow_other):
    if allow_other:
        dv.showErrorMessage = False
    else:
        dv.showErrorMessage = True
        dv.errorTitle = "Invalid value"
        dv.error = "Pick a value from the dropdown list."


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# INSTRUCTIONS sheet
# ---------------------------------------------------------------------------
def build_instructions(wb):
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.sheet_properties.tabColor = "2563EB"
    lines = [
        ("Lineage Canvas — Table Ingestion Template", TITLE_FONT),
        ("", None),
        ("This workbook ingests TABLE-LEVEL and COLUMN-LEVEL metadata only — no actual table data is stored.", None),
        ("", None),
        ("How it works", SECTION_FONT),
        (f"1. There are {MAX_TABLES} fixed table sheets: TABLE_1 .. TABLE_{MAX_TABLES}. Fill one per table you want to ingest (max {MAX_TABLES}).", None),
        ("2. In MASTER, the TABLE REGISTRY assigns a table_name to each sheet you use. ONLY sheets given a table_name there are imported.", None),
        ("3. On each TABLE_n sheet, fill the TABLE METADATA block (namespace, description, ...) and list the columns in the grid below.", None),
        ("   The table's NAME is taken from the registry, so it is not repeated on the sheet. The system (Legacy/Target) is chosen in the app at import time.", None),
        ("4. Cells with a dropdown arrow are restricted to the allowed values. data_type is free text — type any type name.", None),
        ("5. Declare lineage in MASTER: TABLE CONNECTIONS (table→table) and COLUMN CONNECTIONS (column→column). Tables are picked from the registry.", None),
        ("6. Save and upload with the 'Upload Excel' button in the app, then pick the project, canvas and system (Legacy/Target) on the import screen.", None),
        ("", None),
        ("Project / canvas / system", SECTION_FONT),
        ("- You choose the target project, canvas and system (Legacy or Target) in the app's import screen — they are NOT in this workbook.", None),
        ("- A single workbook imports into ONE system. Use a separate import for the other system.", None),
        ("", None),
        ("The MASTER sheet", SECTION_FONT),
        ("- TABLE REGISTRY: sheet_name (pre-filled, do not edit) -> table_name (you fill). Leave table_name blank for sheets you are not using.", None),
        ("- TABLE CONNECTIONS: from_table -> to_table. Pick both from the registry dropdown; optionally add a description.", None),
        ("- COLUMN CONNECTIONS: target column <- source column. Pick the tables from the dropdown and type the column names. Repeat the target to add more sources.", None),
        ("", None),
        ("EXAMPLE_CUSTOMERS is a filled reference sheet — it is NOT imported (it is not in the registry). Use it as a guide.", EXAMPLE_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
    ws.column_dimensions["A"].width = 150
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# MASTER sheet
# ---------------------------------------------------------------------------
def build_master(wb):
    ws = wb.create_sheet("MASTER")
    ws.sheet_properties.tabColor = "047857"
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value="MASTER — table registry & lineage connections").font = TITLE_FONT

    def section_title(r, text):
        ws.cell(row=r, column=1, value=text).font = SECTION_FONT

    def header(r, cols):
        for i, h in enumerate(cols, start=1):
            ws.cell(row=r, column=i, value=h)
        style_header_row(ws, r, len(cols))

    def input_block(r0, r1, n_cols, fill=INPUT_FILL):
        for rr in range(r0, r1 + 1):
            for cc in range(1, n_cols + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = fill
                cell.border = BORDER

    # The project, canvas and system (Legacy/Target) are chosen in the app's
    # import validation screen — not in this workbook. One system per import.
    ws.cell(row=2, column=1,
            value="Project, canvas and system (Legacy/Target) are chosen in the app when you import. One system per import.").font = NOTE_FONT

    # ---- 1) TABLE REGISTRY -------------------------------------------------
    reg_title = 4
    section_title(reg_title, f"1) TABLE REGISTRY — name each sheet you use. Only rows with a table_name are imported (max {MAX_TABLES}).")
    reg_header = reg_title + 1
    header(reg_header, ["sheet_name", "table_name", "Notes"])
    reg_first = reg_header + 1
    reg_last = reg_first + MAX_TABLES - 1
    for i in range(MAX_TABLES):
        rr = reg_first + i
        sc = ws.cell(row=rr, column=1, value=f"TABLE_{i + 1}")   # pre-filled, matches the tab name
        sc.font = KEY_FONT; sc.fill = LOCKED_FILL; sc.border = BORDER
        tc = ws.cell(row=rr, column=2); tc.fill = INPUT_FILL; tc.border = BORDER
        nc = ws.cell(row=rr, column=3); nc.fill = INPUT_FILL; nc.border = BORDER
    ws.cell(row=reg_first, column=3, value="<- type this table's name; leave blank to skip the sheet").font = NOTE_FONT
    registry_range = f"=MASTER!$B${reg_first}:$B${reg_last}"   # table_name column

    # ---- 3) TABLE CONNECTIONS ---------------------------------------------
    tc_title = reg_last + 2
    section_title(tc_title, "2) TABLE CONNECTIONS — table-to-table lineage; one row per edge. Pick tables from the registry.")
    tc_header = tc_title + 1
    header(tc_header, ["from_table", "to_table", "description"])
    tc_first = tc_header + 1
    tc_last = tc_first + 30 - 1
    input_block(tc_first, tc_last, 3)
    dv_from = range_dv(registry_range); ws.add_data_validation(dv_from); dv_from.add(f"A{tc_first}:A{tc_last}")
    dv_to = range_dv(registry_range);   ws.add_data_validation(dv_to);   dv_to.add(f"B{tc_first}:B{tc_last}")

    # ---- 4) COLUMN CONNECTIONS --------------------------------------------
    cc_title = tc_last + 2
    section_title(cc_title, "3) COLUMN CONNECTIONS — column-to-column lineage; one row per source→target mapping. Repeat the target to add more sources.")
    cc_header = cc_title + 1
    header(cc_header, ["target_table", "target_column", "source_table", "source_column"])
    cc_first = cc_header + 1
    cc_last = cc_first + 60 - 1
    input_block(cc_first, cc_last, 4)
    dv_tgt = range_dv(registry_range); ws.add_data_validation(dv_tgt); dv_tgt.add(f"A{cc_first}:A{cc_last}")
    dv_src = range_dv(registry_range); ws.add_data_validation(dv_src); dv_src.add(f"C{cc_first}:C{cc_last}")

    for col, w in {"A": 24, "B": 24, "C": 22, "D": 22}.items():
        ws.column_dimensions[col].width = w
    return ws


# ---------------------------------------------------------------------------
# Table sheet
# ---------------------------------------------------------------------------
def build_table_sheet(wb, name, example=None, title=None):
    """example: optional dict with 'meta' (key->value) and 'columns' (list of dicts)."""
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    if example is not None:
        ws.sheet_properties.tabColor = "10B981"

    ws.cell(row=TABLE_TITLE_ROW, column=1,
            value=title or "TABLE METADATA — this sheet is one table. Its name is set in the MASTER registry.").font = TITLE_FONT

    ws.cell(row=META_HEADER_ROW, column=1, value="Field")
    ws.cell(row=META_HEADER_ROW, column=2, value="Value")
    ws.cell(row=META_HEADER_ROW, column=3, value="Notes")
    style_header_row(ws, META_HEADER_ROW, 3)

    meta_values = (example or {}).get("meta", {}) if example else {}
    row_index = {}
    for offset, (key, note) in enumerate(TABLE_META_FIELDS):
        r = META_FIRST_ROW + offset
        row_index[key] = r
        kc = ws.cell(row=r, column=1, value=key); kc.font = KEY_FONT; kc.fill = KEY_FILL; kc.border = BORDER
        vc = ws.cell(row=r, column=2, value=meta_values.get(key)); vc.fill = INPUT_FILL; vc.border = BORDER
        if example is not None:
            vc.font = EXAMPLE_FONT
        ws.cell(row=r, column=3, value=note).font = NOTE_FONT

    dv_env = list_dv(ENVIRONMENTS); ws.add_data_validation(dv_env); dv_env.add(f"B{row_index['environment']}")
    dv_pk = list_dv(BOOLEANS); ws.add_data_validation(dv_pk); dv_pk.add(f"B{row_index['has_primary_key']}")
    dv_refresh = list_dv(REFRESH); ws.add_data_validation(dv_refresh); dv_refresh.add(f"B{row_index['refresh_frequency']}")

    # Columns section
    ws.cell(row=COLUMNS_TITLE_ROW, column=1, value="COLUMN METADATA — one row per column").font = SECTION_FONT
    for i, h in enumerate(COLUMN_HEADERS, start=1):
        ws.cell(row=COLUMNS_HEADER_ROW, column=i, value=h)
    style_header_row(ws, COLUMNS_HEADER_ROW, len(COLUMN_HEADERS))

    last_data_row = COLUMNS_FIRST_DATA_ROW + COLUMN_DATA_ROWS - 1
    for i in range(len(COLUMN_HEADERS)):
        for r in range(COLUMNS_FIRST_DATA_ROW, last_data_row + 1):
            cell = ws.cell(row=r, column=i + 1); cell.fill = INPUT_FILL; cell.border = BORDER

    for j, colrow in enumerate((example or {}).get("columns", []) if example else []):
        r = COLUMNS_FIRST_DATA_ROW + j
        for i, h in enumerate(COLUMN_HEADERS):
            if h in colrow:
                c = ws.cell(row=r, column=i + 1, value=colrow[h]); c.font = EXAMPLE_FONT

    # data_type is intentionally free text (no dropdown).
    nullable_col = get_column_letter(COLUMN_HEADERS.index("nullable") + 1)
    dv_nullable = list_dv(BOOLEANS); ws.add_data_validation(dv_nullable)
    dv_nullable.add(f"{nullable_col}{COLUMNS_FIRST_DATA_ROW}:{nullable_col}{last_data_row}")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 44
    for i in range(3, len(COLUMN_HEADERS)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 16
    ws.freeze_panes = f"A{COLUMNS_FIRST_DATA_ROW}"
    return ws


EXAMPLE_CUSTOMERS = {
    "meta": {
        "namespace": "ANALYTICS.CORE",
        "description": "One row per customer (cleansed master record).",
        "environment": "PROD",
        "business_domain": "Policy",
        "row_count": 1250000,
        "column_count": 4,
        "has_primary_key": "TRUE",
        "unique_key_columns": "CUSTOMER_ID",
        "grain_description": "one row per customer",
        "refresh_frequency": "DAILY",
    },
    "columns": [
        {"column_name": "CUSTOMER_ID", "data_type": "NUMBER", "nullable": "FALSE",
         "precision": 38, "column_definition": "Surrogate key", "null_count": 0,
         "unique_count": 1250000},
        {"column_name": "FULL_NAME", "data_type": "VARCHAR", "nullable": "FALSE",
         "max_length": 200, "column_definition": "Customer full name"},
        {"column_name": "SIGNUP_DATE", "data_type": "DATE", "nullable": "TRUE",
         "column_definition": "Date the customer first signed up"},
        {"column_name": "LIFETIME_VALUE", "data_type": "DECIMAL", "nullable": "TRUE",
         "precision": 12, "column_computation_formula": "SUM(orders.amount)",
         "min_value": "0", "mean_value": 842.55},
    ],
}


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_instructions(wb)
    build_master(wb)
    build_table_sheet(
        wb, "EXAMPLE_CUSTOMERS", example=EXAMPLE_CUSTOMERS,
        title="EXAMPLE (reference only — NOT imported). Registered in MASTER as table 'CUSTOMERS'.",
    )
    for i in range(1, MAX_TABLES + 1):
        build_table_sheet(wb, f"TABLE_{i}")

    out_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "public", "templates"))
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Lineage_Canvas_Template.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
