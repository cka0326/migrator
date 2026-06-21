#!/usr/bin/env python3
"""
Produce a filled SAMPLE workbook for testing ingestion.

Loads the generated template (preserving its dropdowns/validations and styling)
and fills it with a coherent Property & Casualty (P&C) insurance model so the
importer can be exercised end to end. The workbook imports into ONE system; the
project, canvas and system (Legacy/Target) are chosen in the app at import time.

    /tmp/tplvenv/bin/python scripts/build_sample.py

Output: public/templates/Lineage_Canvas_Sample_Filled.xlsx
"""

import os
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
TEMPLATE = os.path.abspath(os.path.join(HERE, "..", "public", "templates", "Lineage_Canvas_Template.xlsx"))
OUT = os.path.abspath(os.path.join(HERE, "..", "public", "templates", "Lineage_Canvas_Sample_Filled.xlsx"))


def set_kv(ws, key, value):
    """Set the value next to a key in a 'Field | Value' block (column A -> B)."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == key:
            ws.cell(row=row[0].row, column=2, value=value)
            return
    raise KeyError(f"{key} not found on {ws.title}")


def header_row(ws, first_cell):
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip() == first_cell:
            return row[0].row
    raise KeyError(f"header {first_cell} not found on {ws.title}")


def headers_at(ws, hrow):
    out = {}
    c = 1
    while True:
        v = ws.cell(row=hrow, column=c).value
        if v is None or str(v).strip() == "":
            break
        out[str(v).strip()] = c
        c += 1
    return out


def fill_grid(ws, first_cell, rows):
    """Fill rows under a header row, mapping dict keys to header columns."""
    hrow = header_row(ws, first_cell)
    cols = headers_at(ws, hrow)
    for i, record in enumerate(rows):
        r = hrow + 1 + i
        for key, val in record.items():
            if key in cols:
                ws.cell(row=r, column=cols[key], value=val)


def fill_table(ws, meta, columns):
    for k, v in meta.items():
        set_kv(ws, k, v)
    fill_grid(ws, "column_name", columns)


def fill_registry(ws, mapping):
    """mapping: {sheet_name: table_name}. Writes table_name beside each sheet row."""
    hrow = header_row(ws, "sheet_name")
    r = hrow + 1
    while True:
        sheet = ws.cell(row=r, column=1).value
        if sheet is None or str(sheet).strip() == "":
            break
        name = mapping.get(str(sheet).strip())
        if name:
            ws.cell(row=r, column=2, value=name)
        r += 1


# ---------------------------------------------------------------------------
# Sample data — Property & Casualty insurance model (single system)
# ---------------------------------------------------------------------------
REGISTRY = {
    "TABLE_1": "PARTY",
    "TABLE_2": "POLICY",
    "TABLE_3": "COVERAGE",
    "TABLE_4": "INSURED_LOCATION",
    "TABLE_5": "CLAIM",
    "TABLE_6": "CLAIM_TRANSACTION",
    "TABLE_7": "POLICY_PREMIUM_SUMMARY",
}

TABLES = {
    # 1) PARTY — policyholders / insured parties
    "TABLE_1": (
        {
            "namespace": "STG.INSURANCE",
            "description": "Insured parties / policyholders (person or organization).",
            "environment": "PROD", "business_domain": "Party",
            "row_count": 920000, "has_primary_key": "TRUE",
            "unique_key_columns": "PARTY_ID", "grain_description": "one row per party",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "PARTY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "null_count": 0, "unique_count": 920000, "column_definition": "Surrogate key for a party"},
            {"column_name": "PARTY_NAME", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 200},
            {"column_name": "PARTY_TYPE", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 20, "column_definition": "PERSON or ORGANIZATION"},
            {"column_name": "DATE_OF_BIRTH", "data_type": "DATE", "nullable": "TRUE"},
            {"column_name": "TAX_ID", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 20},
        ],
    ),
    # 2) POLICY — policy contracts
    "TABLE_2": (
        {
            "namespace": "CORE.INSURANCE",
            "description": "P&C policy contracts (one row per policy term).",
            "environment": "PROD", "business_domain": "Policy",
            "row_count": 1450000, "has_primary_key": "TRUE",
            "unique_key_columns": "POLICY_ID", "grain_description": "one row per policy per term",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "POLICY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "null_count": 0, "unique_count": 1450000, "column_definition": "Surrogate key for a policy term"},
            {"column_name": "POLICY_NUMBER", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 30},
            {"column_name": "PARTY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "column_definition": "FK to PARTY (named insured)"},
            {"column_name": "PRODUCT_CODE", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 20, "column_definition": "HOME / AUTO / CPP / BOP ..."},
            {"column_name": "EFFECTIVE_DATE", "data_type": "DATE", "nullable": "FALSE"},
            {"column_name": "EXPIRATION_DATE", "data_type": "DATE", "nullable": "FALSE"},
            {"column_name": "POLICY_STATUS", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 20, "column_definition": "QUOTED / BOUND / INFORCE / CANCELLED / EXPIRED"},
            {"column_name": "WRITTEN_PREMIUM", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 14, "min_value": "0", "mean_value": 1284.50},
        ],
    ),
    # 3) COVERAGE — coverages attached to a policy
    "TABLE_3": (
        {
            "namespace": "CORE.INSURANCE",
            "description": "Coverages attached to a policy (limits / deductibles / premium).",
            "environment": "PROD", "business_domain": "Policy",
            "row_count": 5200000, "has_primary_key": "TRUE",
            "unique_key_columns": "COVERAGE_ID", "grain_description": "one row per coverage per policy",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "COVERAGE_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "unique_count": 5200000},
            {"column_name": "POLICY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "column_definition": "FK to POLICY"},
            {"column_name": "COVERAGE_CODE", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 20, "column_definition": "DWELLING / LIABILITY / COLLISION / COMPREHENSIVE ..."},
            {"column_name": "LIMIT_AMOUNT", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16},
            {"column_name": "DEDUCTIBLE_AMOUNT", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 14},
            {"column_name": "COVERAGE_PREMIUM", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 14, "min_value": "0", "mean_value": 357.80},
        ],
    ),
    # 4) INSURED_LOCATION — property risk locations
    "TABLE_4": (
        {
            "namespace": "CORE.INSURANCE",
            "description": "Insured property locations (risk addresses).",
            "environment": "PROD", "business_domain": "Risk",
            "row_count": 1610000, "has_primary_key": "TRUE",
            "unique_key_columns": "LOCATION_ID", "grain_description": "one row per insured location",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "LOCATION_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "unique_count": 1610000},
            {"column_name": "POLICY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "column_definition": "FK to POLICY"},
            {"column_name": "ADDRESS_LINE1", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 200},
            {"column_name": "CITY", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 100},
            {"column_name": "STATE", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 2},
            {"column_name": "POSTAL_CODE", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 10},
            {"column_name": "CONSTRUCTION_TYPE", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 30, "column_definition": "FRAME / MASONRY / FIRE_RESISTIVE ..."},
            {"column_name": "YEAR_BUILT", "data_type": "NUMBER", "nullable": "TRUE", "precision": 4},
            {"column_name": "TIV", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16, "column_definition": "Total insured value at this location", "min_value": "0"},
        ],
    ),
    # 5) CLAIM — claims raised against policies
    "TABLE_5": (
        {
            "namespace": "CORE.INSURANCE",
            "description": "Claims raised against policies/coverages.",
            "environment": "PROD", "business_domain": "Claims",
            "row_count": 410000, "has_primary_key": "TRUE",
            "unique_key_columns": "CLAIM_ID", "grain_description": "one row per claim",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "CLAIM_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "unique_count": 410000},
            {"column_name": "CLAIM_NUMBER", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 30},
            {"column_name": "POLICY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "column_definition": "FK to POLICY"},
            {"column_name": "COVERAGE_ID", "data_type": "NUMBER", "nullable": "TRUE", "precision": 38, "column_definition": "FK to COVERAGE"},
            {"column_name": "LOSS_DATE", "data_type": "DATE", "nullable": "FALSE"},
            {"column_name": "REPORT_DATE", "data_type": "DATE", "nullable": "TRUE"},
            {"column_name": "CLAIM_STATUS", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 20, "column_definition": "OPEN / CLOSED / REOPENED / DENIED"},
            {"column_name": "CAUSE_OF_LOSS", "data_type": "VARCHAR", "nullable": "TRUE", "max_length": 40, "column_definition": "FIRE / WATER / THEFT / COLLISION / WIND ..."},
            {"column_name": "RESERVE_AMOUNT", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16, "min_value": "0"},
        ],
    ),
    # 6) CLAIM_TRANSACTION — financial transactions on a claim
    "TABLE_6": (
        {
            "namespace": "CORE.INSURANCE",
            "description": "Financial transactions on a claim (payments and reserve changes).",
            "environment": "PROD", "business_domain": "Claims",
            "row_count": 2750000, "has_primary_key": "TRUE",
            "unique_key_columns": "CLAIM_TXN_ID", "grain_description": "one row per claim financial transaction",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "CLAIM_TXN_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "unique_count": 2750000},
            {"column_name": "CLAIM_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38, "column_definition": "FK to CLAIM"},
            {"column_name": "TXN_TYPE", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 20, "column_definition": "PAYMENT / RESERVE / RECOVERY / EXPENSE"},
            {"column_name": "TXN_DATE", "data_type": "DATE", "nullable": "FALSE"},
            {"column_name": "PAID_AMOUNT", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16, "min_value": "0", "mean_value": 3120.44},
            {"column_name": "RESERVE_CHANGE", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16},
        ],
    ),
    # 7) POLICY_PREMIUM_SUMMARY — analytic rollup
    "TABLE_7": (
        {
            "namespace": "MART.INSURANCE",
            "description": "Per-policy rollup of premium, exposure and claims for reporting.",
            "environment": "PROD", "business_domain": "Finance",
            "row_count": 1450000, "has_primary_key": "TRUE",
            "unique_key_columns": "POLICY_ID", "grain_description": "one row per policy",
            "refresh_frequency": "DAILY",
        },
        [
            {"column_name": "POLICY_ID", "data_type": "NUMBER", "nullable": "FALSE", "precision": 38},
            {"column_name": "POLICY_NUMBER", "data_type": "VARCHAR", "nullable": "FALSE", "max_length": 30},
            {"column_name": "TOTAL_WRITTEN_PREMIUM", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16, "column_computation_formula": "SUM(POLICY.WRITTEN_PREMIUM)"},
            {"column_name": "TOTAL_COVERAGE_PREMIUM", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 16, "column_computation_formula": "SUM(COVERAGE.COVERAGE_PREMIUM)"},
            {"column_name": "TOTAL_TIV", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 18, "column_computation_formula": "SUM(INSURED_LOCATION.TIV)"},
            {"column_name": "CLAIM_COUNT", "data_type": "NUMBER", "nullable": "TRUE", "precision": 10, "column_computation_formula": "COUNT(DISTINCT CLAIM.CLAIM_ID)"},
            {"column_name": "TOTAL_INCURRED", "data_type": "DECIMAL", "nullable": "TRUE", "precision": 18, "column_computation_formula": "SUM(CLAIM_TRANSACTION.PAID_AMOUNT + CLAIM_TRANSACTION.RESERVE_CHANGE)"},
        ],
    ),
}

TABLE_CONNECTIONS = [
    {"from_table": "PARTY", "to_table": "POLICY", "description": "Named insured feeds the policy."},
    {"from_table": "POLICY", "to_table": "COVERAGE", "description": "A policy has many coverages."},
    {"from_table": "POLICY", "to_table": "INSURED_LOCATION", "description": "A policy insures one or more locations."},
    {"from_table": "POLICY", "to_table": "CLAIM", "description": "Claims are made against a policy."},
    {"from_table": "COVERAGE", "to_table": "CLAIM", "description": "A claim is tied to a coverage."},
    {"from_table": "CLAIM", "to_table": "CLAIM_TRANSACTION", "description": "Claims have financial transactions."},
    {"from_table": "POLICY", "to_table": "POLICY_PREMIUM_SUMMARY", "description": "Policy attributes feed the rollup."},
    {"from_table": "COVERAGE", "to_table": "POLICY_PREMIUM_SUMMARY", "description": "Coverage premium aggregated."},
    {"from_table": "INSURED_LOCATION", "to_table": "POLICY_PREMIUM_SUMMARY", "description": "TIV aggregated."},
    {"from_table": "CLAIM_TRANSACTION", "to_table": "POLICY_PREMIUM_SUMMARY", "description": "Incurred losses aggregated."},
]

COLUMN_CONNECTIONS = [
    {"target_table": "POLICY", "target_column": "PARTY_ID", "source_table": "PARTY", "source_column": "PARTY_ID"},
    {"target_table": "COVERAGE", "target_column": "POLICY_ID", "source_table": "POLICY", "source_column": "POLICY_ID"},
    {"target_table": "INSURED_LOCATION", "target_column": "POLICY_ID", "source_table": "POLICY", "source_column": "POLICY_ID"},
    {"target_table": "CLAIM", "target_column": "POLICY_ID", "source_table": "POLICY", "source_column": "POLICY_ID"},
    {"target_table": "CLAIM", "target_column": "COVERAGE_ID", "source_table": "COVERAGE", "source_column": "COVERAGE_ID"},
    {"target_table": "CLAIM_TRANSACTION", "target_column": "CLAIM_ID", "source_table": "CLAIM", "source_column": "CLAIM_ID"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "POLICY_ID", "source_table": "POLICY", "source_column": "POLICY_ID"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "POLICY_NUMBER", "source_table": "POLICY", "source_column": "POLICY_NUMBER"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "TOTAL_WRITTEN_PREMIUM", "source_table": "POLICY", "source_column": "WRITTEN_PREMIUM"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "TOTAL_COVERAGE_PREMIUM", "source_table": "COVERAGE", "source_column": "COVERAGE_PREMIUM"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "TOTAL_TIV", "source_table": "INSURED_LOCATION", "source_column": "TIV"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "CLAIM_COUNT", "source_table": "CLAIM", "source_column": "CLAIM_ID"},
    # Two sources for one target column (repeat the target):
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "TOTAL_INCURRED", "source_table": "CLAIM_TRANSACTION", "source_column": "PAID_AMOUNT"},
    {"target_table": "POLICY_PREMIUM_SUMMARY", "target_column": "TOTAL_INCURRED", "source_table": "CLAIM_TRANSACTION", "source_column": "RESERVE_CHANGE"},
]


def main():
    wb = load_workbook(TEMPLATE)

    master = wb["MASTER"]
    fill_registry(master, REGISTRY)
    fill_grid(master, "from_table", TABLE_CONNECTIONS)
    fill_grid(master, "target_table", COLUMN_CONNECTIONS)

    for sheet, (meta, columns) in TABLES.items():
        fill_table(wb[sheet], meta, columns)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
